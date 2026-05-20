"""
Multi-sheet Excel export of the three financial statements.

Layout per workbook:
- Sheet "Income Statement"     — INCOME_STATEMENT_ORDER, one column per period
- Sheet "Balance Sheet"        — BALANCE_SHEET_ORDER, with section headers
- Sheet "Cash Flow"            — CASH_FLOW_ORDER

Number cells are real numbers (not strings) so the user can pivot or
chart inside Excel. The displayed format is ``"#,##0.00"`` which Excel
applies on the fly. Headers / section rows are formatted in the brand
gold + dark-navy palette via openpyxl ``Font`` + ``PatternFill``.

Returns ``bytes`` so the page can wire it directly into
``st.download_button(data=…, file_name="…")``.
"""
from __future__ import annotations
from io import BytesIO
from typing import Optional

import pandas as pd

from core.account_labels import (
    INCOME_STATEMENT_ORDER, BALANCE_SHEET_ORDER, CASH_FLOW_ORDER,
    SECTION_LABELS, get_label,
)
from core.formatters import format_period


# ============================================================
# Style constants — match the Streamlit theme
# ============================================================
_BRAND_GOLD = "C9A961"
_DARK_NAVY  = "0B0E14"
_PANEL_BG   = "131826"
_BORDER     = "1F2937"
_TEXT_PRIM  = "E8EAED"
_TEXT_MUTED = "9CA3AF"
_GAINS      = "10B981"
_LOSSES     = "EF4444"


def _style_header(cell, *, gold: bool = False) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    cell.font = Font(name="Calibri", size=11, bold=True,
                     color=(_BRAND_GOLD if gold else _TEXT_MUTED))
    cell.fill = PatternFill("solid", fgColor=_PANEL_BG)
    cell.alignment = Alignment(horizontal="right" if not gold else "left",
                               vertical="center")


def _style_section(cell) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    cell.font = Font(name="Calibri", size=11, bold=True, color=_BRAND_GOLD)
    cell.fill = PatternFill("solid", fgColor=_PANEL_BG)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def _style_value(cell, *, subtotal: bool = False) -> None:
    from openpyxl.styles import Alignment, Font
    cell.font = Font(
        name="Calibri", size=11,
        bold=subtotal, color=_TEXT_PRIM,
    )
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'


def _style_label(cell, *, subtotal: bool = False) -> None:
    from openpyxl.styles import Alignment, Font
    cell.font = Font(
        name="Calibri", size=11,
        bold=subtotal, color=_TEXT_PRIM,
    )
    cell.alignment = Alignment(horizontal="left", vertical="center")


# ============================================================
# Single-sheet writer
# ============================================================
def _write_statement(
    ws,
    df: pd.DataFrame,
    order: list[tuple[str, str]],
    *,
    sheet_label: str,
) -> None:
    """Write one statement's rows + header into the worksheet."""
    if df is None or df.empty:
        ws["A1"] = f"No data for {sheet_label}"
        return

    df = df.sort_index()
    periods = list(df.index)

    # Header row 1: ($USD)  ·  FY 2019  ·  FY 2020 …
    ws.cell(row=1, column=1, value=f"{sheet_label}  ($USD)")
    _style_header(ws.cell(row=1, column=1), gold=True)
    for col_idx, period in enumerate(periods, start=2):
        ws.cell(row=1, column=col_idx, value=format_period(period))
        _style_header(ws.cell(row=1, column=col_idx))

    # Body rows
    row_idx = 2
    for key, kind in order:
        if kind == "section":
            label = SECTION_LABELS.get(key, key)
            ws.cell(row=row_idx, column=1, value=label)
            _style_section(ws.cell(row=row_idx, column=1))
            for c in range(2, 2 + len(periods)):
                _style_section(ws.cell(row=row_idx, column=c))
            row_idx += 1
            continue

        if key not in df.columns:
            continue

        is_subtotal = (kind == "subtotal")
        ws.cell(row=row_idx, column=1, value=get_label(key))
        _style_label(ws.cell(row=row_idx, column=1), subtotal=is_subtotal)
        for col_idx, period in enumerate(periods, start=2):
            try:
                v = df.loc[period, key]
                if pd.isna(v):
                    raise ValueError("nan")
                ws.cell(row=row_idx, column=col_idx, value=float(v))
            except Exception:
                ws.cell(row=row_idx, column=col_idx, value=None)
            _style_value(ws.cell(row=row_idx, column=col_idx),
                         subtotal=is_subtotal)
        row_idx += 1

    # Auto-fit-ish: name col wider, period cols medium
    ws.column_dimensions["A"].width = 36
    for col_idx in range(2, 2 + len(periods)):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 16


# ============================================================
# Public API
# ============================================================
def export_financials_xlsx(
    *,
    income: Optional[pd.DataFrame],
    balance: Optional[pd.DataFrame],
    cash: Optional[pd.DataFrame],
    ticker: str = "TICKER",
) -> bytes:
    """Build the workbook in memory and return its bytes."""
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise ImportError(
            "openpyxl is required for Excel export. "
            "Install with `pip install openpyxl`."
        ) from exc

    wb = Workbook()
    # Default first sheet
    ws_inc = wb.active
    ws_inc.title = "Income Statement"
    _write_statement(ws_inc, income, INCOME_STATEMENT_ORDER,
                     sheet_label=f"{ticker} · Income")
    ws_bal = wb.create_sheet("Balance Sheet")
    _write_statement(ws_bal, balance, BALANCE_SHEET_ORDER,
                     sheet_label=f"{ticker} · Balance")
    ws_cf = wb.create_sheet("Cash Flow")
    _write_statement(ws_cf, cash, CASH_FLOW_ORDER,
                     sheet_label=f"{ticker} · Cash Flow")

    out = BytesIO()
    wb.save(out)
    return out.getvalue()
